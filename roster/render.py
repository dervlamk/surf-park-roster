"""
Render rotation slots as an xlsx workbook structured like the reference
LG Schedule sheet in parameters/spec_surf_park_roster_v0.xlsx.

Per-sheet layout:
  Row 1: section banner ("Shore Guards" / "Reef Guards" / …) merged across
         that section's columns.
  Row 2: per-staff column header (SG1, SG2, …).
  Row 3+: one row per 30-min slot, with time + state + reef/bay wave +
         each staff member's station for that slot.

Public API:
    render_xlsx(slots_by_day, out_path)  — write one sheet per day.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from roster.rotation import Slot, staff_layout

# ── light colour palette (no domain meaning, just visual grouping) ───────────
_SECTION_FILL = {
    "Shore Guards":     PatternFill("solid", fgColor="FFE3F2FD"),
    "Reef Guards":      PatternFill("solid", fgColor="FFFFF3E0"),
    "Advanced Coaches": PatternFill("solid", fgColor="FFE8F5E9"),
    "Bay Coaches":      PatternFill("solid", fgColor="FFF3E5F5"),
}
_HEADER_FONT = Font(bold=True)
_OFF_FILL    = PatternFill("solid", fgColor="FFEEEEEE")


def render_xlsx(
    slots_by_day: dict[str, list[Slot]],
    out_path: str | Path,
) -> Path:
    """
    Write one sheet per day to a single workbook.  ``slots_by_day`` is an
    insertion-ordered mapping from sheet name (e.g. "Monday") to its slot
    list.  Empty values are skipped.
    """
    wb = openpyxl.Workbook()
    # Drop the default empty sheet that openpyxl creates.
    wb.remove(wb.active)
    for sheet_name, slots in slots_by_day.items():
        if not slots:
            continue
        _add_sheet(wb, sheet_name, slots)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _add_sheet(wb, sheet_name: str, slots: list[Slot]) -> None:
    layout = staff_layout()
    ws = wb.create_sheet(title=sheet_name)

    # Fixed leading columns: time, state, reef wave, bay wave
    lead = ["Time", "State", "Reef Wave", "Bay Wave"]
    n_lead = len(lead)

    # Build per-section column ranges
    col_cursor = n_lead + 1
    section_ranges: list[tuple[str, int, int, list[str]]] = []  # (name, start_col, end_col, ids)
    for section, ids in layout.items():
        start = col_cursor
        end = col_cursor + len(ids) - 1
        section_ranges.append((section, start, end, ids))
        col_cursor = end + 1

    # Row 1: section banners (merged across each section)
    for col, label in enumerate(lead, start=1):
        c = ws.cell(row=1, column=col, value="")
    for name, start, end, _ids in section_ranges:
        c = ws.cell(row=1, column=start, value=name)
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.fill = _SECTION_FILL[name]
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    # Row 2: column headers
    for col, label in enumerate(lead, start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = _HEADER_FONT
    for _name, start, _end, ids in section_ranges:
        for offset, sid in enumerate(ids):
            c = ws.cell(row=2, column=start + offset, value=sid)
            c.font = _HEADER_FONT
            c.alignment = Alignment(horizontal="center")

    # Data rows
    for r, slot in enumerate(slots, start=3):
        ws.cell(row=r, column=1, value=slot.wallclock)
        ws.cell(row=r, column=2, value=slot.state)
        ws.cell(row=r, column=3, value=slot.reef_wave or "")
        ws.cell(row=r, column=4, value=slot.bay_wave or "")
        for _name, start, _end, ids in section_ranges:
            for offset, sid in enumerate(ids):
                val = slot.assignments.get(sid, "")
                c = ws.cell(row=r, column=start + offset, value=val)
                c.alignment = Alignment(horizontal="center")
                if val == "OFF":
                    c.fill = _OFF_FILL

    # Column widths
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 26
    for _name, start, end, _ids in section_ranges:
        for col in range(start, end + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    # Freeze header rows + leading columns
    ws.freeze_panes = "E3"
