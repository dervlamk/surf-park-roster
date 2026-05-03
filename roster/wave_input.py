"""
Load a wave program (one column) from the OPENING SCHEDULE FINAL 1.0 sheet.

The sheet has TWO side-by-side hourly tables:
  * Reef waves: column A = time, columns B..H = Mon..Sun
  * Bay waves:  column K = time, columns L..R = Mon..Sun

Plus a summary block at the bottom of the sheet (currently row 43 = "Open
Time", row 44 = "Open Hours", row 45 = "Wave Hours") that gives each day's
operating window in shorthand like ``7-8:30`` (open hour AM, close HH:MM PM).

Public API:
    load_wave_program(...)  → {hour_int: HourProgram}        (per-hour wave types)
    load_open_window(...)   → (open_hhmm, close_hhmm)        (e.g. "07:00", "20:30")
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import time as _time
from pathlib import Path

import openpyxl


@dataclass(frozen=True)
class HourProgram:
    hour: int            # 0..23
    reef_wave: str | None
    bay_wave: str | None


def _parse_hour(raw) -> int | None:
    """Coerce a time-column cell into an integer hour-of-day, or None."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, _time):
        return raw.hour
    if isinstance(raw, str):
        s = raw.strip()
        # Strip trailing notes like "7:00:00 PM\nBarrel Hour" → "7:00:00 PM"
        s = s.split("\n", 1)[0].strip()
        # Try "HH:MM:SS" / "H:MM:SS" / "H:MM:SS PM"
        ampm = None
        if s.lower().endswith(" am") or s.lower().endswith(" pm"):
            ampm = s[-2:].lower()
            s = s[:-3].strip()
        try:
            parts = [int(x) for x in s.split(":")]
        except ValueError:
            return None
        h = parts[0]
        if ampm == "pm" and h < 12:
            h += 12
        if ampm == "am" and h == 12:
            h = 0
        return h
    return None


def _normalize_label(raw) -> str | None:
    """Cells contain extra whitespace and the literal '0' for closed slots."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("", "0"):
        return None
    # collapse runs of whitespace introduced by Excel cell wrapping
    return " ".join(s.split())


def load_wave_program(
    xlsx_path: str | Path,
    sheet_name: str,
    column_name: str,
) -> dict[int, HourProgram]:
    """
    Read both reef and bay wave-type columns for `column_name`.

    Returns {hour_int: HourProgram}.  Hours not listed in the sheet (or whose
    cells are blank/'0') are omitted, meaning "no wave running this hour".
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {xlsx_path}")
    ws = wb[sheet_name]

    # The sheet has two parallel tables sharing the same row layout.  We
    # use column A as the canonical time index for both — the bay table's
    # own time column (K) has at least one corrupted cell ("Shredder
    # Sessions" pasted into K27 in place of 16:00) which would otherwise
    # silently drop that hour's bay data.
    time_col, reef_data_col = _locate_column(ws, "REEF", column_name)
    _,        bay_data_col  = _locate_column(ws, "BAY",  column_name)

    out: dict[int, HourProgram] = {}
    for row in range(1, ws.max_row + 1):
        h = _parse_hour(ws.cell(row=row, column=time_col).value)
        if h is None:
            continue
        reef = _normalize_label(ws.cell(row=row, column=reef_data_col).value)
        bay  = _normalize_label(ws.cell(row=row, column=bay_data_col).value)
        out[h] = HourProgram(h, reef, bay)

    return out


def _locate_column(ws, side: str, day_header: str) -> tuple[int, int]:
    """
    Locate the (time_col, data_col) pair for the requested side ("REEF" or
    "BAY") and day header (e.g. "Monday").

    Strategy: row 5 holds a banner like "PARK SURF OPENING SCHEDULE REEF
    WAVES PROPOSAL 1.2" / "PARK SURF OPENING SCHEDULE BAY WAVES PROPOSAL 1.2".
    Use it to identify which contiguous block of columns belongs to which
    side, then row 6 holds the day-name headers within each block.
    """
    side_up = side.upper()

    # Step 1: find the column where the side's banner starts.
    banner_col: int | None = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=5, column=col).value
        if v and side_up in str(v).upper() and "WAVES" in str(v).upper():
            banner_col = col
            break
    if banner_col is None:
        raise ValueError(f"could not find {side} banner in row 5")

    # Step 2: within row 6, find the day-name header on or after banner_col.
    target = day_header.strip().lower()
    data_col: int | None = None
    for col in range(banner_col, ws.max_column + 1):
        v = ws.cell(row=6, column=col).value
        if v and str(v).strip().lower() == target:
            data_col = col
            break
    if data_col is None:
        raise ValueError(
            f"could not find day header {day_header!r} on the {side} side"
        )

    # Step 3: time column is the first non-empty header column ≤ data_col
    # within this side's block.  In practice it's the banner column itself.
    return banner_col, data_col


def load_open_window(
    xlsx_path: str | Path,
    sheet_name: str,
    column_name: str,
) -> tuple[str, str]:
    """
    Read the ``Open Time`` summary cell for ``column_name`` and return
    (open_hhmm, close_hhmm) as 24-h "HH:MM" strings.

    The cell format is e.g. ``7-8:30`` meaning open at 7:00 (AM) and close
    at 8:30 PM.  Convention: the left value is morning (taken as-is); the
    right value is treated as PM whenever its hour is ≤ 11.  Hours already
    in 24-h form (12–23) pass through unchanged.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {xlsx_path}")
    ws = wb[sheet_name]

    # Find the summary row whose col-A label equals "Open Time".
    open_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip().lower() == "open time":
            open_row = r
            break
    if open_row is None:
        raise ValueError("could not find 'Open Time' summary row in sheet")

    # Day column is the same column the wave-program loader picks: locate it
    # under the REEF banner (the BAY banner mirrors it identically per row 43).
    _, data_col = _locate_column(ws, "REEF", column_name)
    raw = ws.cell(row=open_row, column=data_col).value
    if raw is None or str(raw).strip() == "":
        raise ValueError(f"no Open Time value for column {column_name!r}")

    return _parse_open_close(str(raw))


def _parse_open_close(raw: str) -> tuple[str, str]:
    """Parse e.g. '7-8:30' → ('07:00', '20:30')."""
    s = raw.strip()
    if "-" not in s:
        raise ValueError(f"open/close cell missing '-': {raw!r}")
    left, right = (p.strip() for p in s.split("-", 1))

    def _hm(part: str) -> tuple[int, int]:
        if ":" in part:
            h, m = part.split(":", 1)
            return int(h), int(m)
        return int(part), 0

    oh, om = _hm(left)
    ch, cm = _hm(right)
    # Right side is evening unless explicitly already 24-h (≥ 12).
    if ch <= 11:
        ch += 12
    return f"{oh:02d}:{om:02d}", f"{ch:02d}:{cm:02d}"


if __name__ == "__main__":
    # Smoke test: dump program + open window for every weekday.
    import config_v2 as cfg

    for day in cfg.DAYS:
        try:
            o, c = load_open_window(cfg.WAVE_INPUT_XLSX, cfg.WAVE_INPUT_SHEET, day)
        except Exception as e:
            print(f"{day:10s}  open window: ERROR {e}")
            continue
        prog = load_wave_program(cfg.WAVE_INPUT_XLSX, cfg.WAVE_INPUT_SHEET, day)
        print(f"{day:10s}  open={o}  close={c}  ({len(prog)} wave hours)")
