"""
Render staffing results as xlsx workbooks.

Produces two files:
  staffing_v0.xlsx  — shift plan + CA notes + FT/PT person-hours breakdown
  roster_v0.xlsx    — individual position assignments per day + shift template table
"""
from __future__ import annotations

from pathlib import Path
from itertools import groupby

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from solver.staffing import DayStaffingResult, ShiftBlock, PersonTypeBreakdown
import config_v2 as cfg


# ── Palette ───────────────────────────────────────────────────────────────────
_BOLD  = Font(bold=True)
_SZ13  = Font(bold=True, size=13)
_GRAY  = PatternFill("solid", fgColor="FFEEEEEE")
_BLUE  = PatternFill("solid", fgColor="FFE3F2FD")
_AMBER = PatternFill("solid", fgColor="FFFFF8E1")
_GREEN = PatternFill("solid", fgColor="FFE8F5E9")
_LILAC = PatternFill("solid", fgColor="FFF3E5F5")
_TEAL  = PatternFill("solid", fgColor="FFE0F7FA")
_ROSE  = PatternFill("solid", fgColor="FFFCE4EC")
_WRAP  = Alignment(wrap_text=True)
_CTR   = Alignment(horizontal="center")
_WCTR  = Alignment(wrap_text=True, horizontal="center")

_ROLE_FILL = {
    "Shore Guard":    _BLUE,
    "Reef Guard":     _AMBER,
    "Advanced Coach": _GREEN,
    "Bay Coach":      _LILAC,
}


def _role_fill(role: str) -> PatternFill:
    for key, fill in _ROLE_FILL.items():
        if role.startswith(key):
            return fill
    return _GRAY


def _header_row(ws, texts: list, fills: list[PatternFill] | None = None) -> None:
    ws.append(texts)
    row = ws.max_row
    for col, txt in enumerate(texts, 1):
        c = ws.cell(row, col)
        c.font = _BOLD
        c.fill = (fills[col - 1] if fills else _GRAY)
        c.alignment = _WCTR


def _bold_row(ws, texts: list) -> None:
    ws.append(texts)
    for col in range(1, len(texts) + 1):
        ws.cell(ws.max_row, col).font = _BOLD


# ── staffing_v0.xlsx — per-day sheet ─────────────────────────────────────────

def _add_day_sheet(wb, result: DayStaffingResult) -> None:
    ws = wb.create_sheet(title=result.day)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.append([f"{result.day}"])
    ws.cell(ws.max_row, 1).font = _SZ13

    ws.append([
        f"Park: {result.park_open_time}–{result.park_close_time}   "
        f"LG shifts: {result.lg_start_time}–{result.lg_end_time}   "
        f"(LG start = 30 min before first wave)"
    ])
    ws.append([])

    # ── Shift plan ────────────────────────────────────────────────────────────
    ws.append(["SHIFT PLAN"])
    ws.cell(ws.max_row, 1).font = _BOLD

    _header_row(ws, ["Role", "Shift Start", "Shift End", "Workers", "Hours", "CA Break Requirements"])

    for sb in result.shifts:
        ws.append([sb.role, sb.start, sb.end, sb.workers, sb.shift_hrs, sb.ca_note or "—"])
        fill = _role_fill(sb.role)
        for col in range(1, 7):
            ws.cell(ws.max_row, col).fill = fill
        ws.cell(ws.max_row, 6).alignment = _WRAP

    ws.append([])

    # ── All-PT person-hours breakdown ─────────────────────────────────────────
    ws.append(["ALL PART-TIME SCENARIO  (every worker does one shift)"])
    ws.cell(ws.max_row, 1).font = _BOLD

    _header_row(ws, ["Role", "Shift Start", "Shift End", "Workers", "Hours/Person", "FT-Eligible?"])

    for b in sorted(result.person_breakdown, key=lambda x: (x.role, x.shift_start)):
        ws.append([
            b.role, b.shift_start, b.shift_end, b.n_workers,
            b.shift_hrs,
            f"Yes (≥{cfg.FT_SHIFT_THRESHOLD:.0f}h)" if b.is_ft_eligible else "No",
        ])
        fill = _role_fill(b.role)
        for col in range(1, 7):
            ws.cell(ws.max_row, col).fill = fill

    ws.append([])
    pt_total = result.min_persons_all_pt
    _bold_row(ws, ["Total unique persons (all PT)", pt_total, "", "", "", ""])
    ws.append([
        "",
        "Each person works exactly one shift. Person count = sum of workers "
        "across all shift blocks shown above.",
    ])
    ws.cell(ws.max_row, 2).alignment = _WRAP
    ws.append([])

    # ── FT mix breakdown ──────────────────────────────────────────────────────
    ws.append([f"FULL-TIME MIX SCENARIO  (shifts ≥ {cfg.FT_SHIFT_THRESHOLD:.0f}h classified as FT-eligible)"])
    ws.cell(ws.max_row, 1).font = _BOLD

    ft_blocks = [b for b in result.person_breakdown if b.is_ft_eligible]
    pt_blocks  = [b for b in result.person_breakdown if not b.is_ft_eligible]

    if ft_blocks:
        ws.append(["FT-eligible shifts (≥ {:.0f}h):".format(cfg.FT_SHIFT_THRESHOLD)])
        _header_row(ws, ["Role", "Shift Start", "Shift End", "Workers", "Hours/Person", ""])
        for b in sorted(ft_blocks, key=lambda x: (x.role, x.shift_start)):
            ws.append([b.role, b.shift_start, b.shift_end, b.n_workers, b.shift_hrs, ""])
            fill = _role_fill(b.role)
            for col in range(1, 6):
                ws.cell(ws.max_row, col).fill = fill

    if pt_blocks:
        ws.append([])
        ws.append(["PT shifts (< {:.0f}h):".format(cfg.FT_SHIFT_THRESHOLD)])
        _header_row(ws, ["Role", "Shift Start", "Shift End", "Workers", "Hours/Person", ""])
        for b in sorted(pt_blocks, key=lambda x: (x.role, x.shift_start)):
            ws.append([b.role, b.shift_start, b.shift_end, b.n_workers, b.shift_hrs, ""])
            fill = _role_fill(b.role)
            for col in range(1, 6):
                ws.cell(ws.max_row, col).fill = fill

    ws.append([])
    _bold_row(ws, ["FT-eligible workers", result.ft_eligible_workers,
                   "PT workers", result.pt_only_workers, "", ""])
    ws.append([
        "",
        f"FT threshold: ≥ {cfg.FT_SHIFT_THRESHOLD:.0f}h/shift. "
        "FT workers are those in shifts at or above this length. "
        f"CA law: ≥ 30 h/week for 90 consecutive days triggers FT offer "
        f"({cfg.FT_SHIFT_THRESHOLD:.0f}h/day × 5 days = "
        f"{cfg.FT_SHIFT_THRESHOLD * 5:.0f}h/week — exactly at threshold).",
    ])
    ws.cell(ws.max_row, 2).alignment = _WRAP
    ws.append([])

    # ── Headcount summary ─────────────────────────────────────────────────────
    ws.append(["HEADCOUNT SUMMARY"])
    ws.cell(ws.max_row, 1).font = _BOLD

    for label, val in [
        ("Shore Guard person-shifts",    result.shore_total),
        ("Reef Guard person-shifts",     result.reef_total),
        ("Advanced Coach shifts (fixed)",result.ac_total),
        ("Bay Coach person-shifts",      result.bay_coach_total),
        ("Total unique persons (all PT)", result.min_persons_all_pt),
    ]:
        ws.append([label, val])
        if "Total" in label:
            ws.cell(ws.max_row, 1).font = _BOLD
            ws.cell(ws.max_row, 2).font = _BOLD

    ws.append([])

    # ── Notes ─────────────────────────────────────────────────────────────────
    if result.notes:
        ws.append(["NOTES"])
        ws.cell(ws.max_row, 1).font = _BOLD
        for note in result.notes:
            ws.append(["", note])
            ws.cell(ws.max_row, 2).alignment = _WRAP

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 55


# ── staffing_v0.xlsx — weekly summary sheet ───────────────────────────────────

def _add_weekly_summary(wb, results: list[DayStaffingResult]) -> None:
    ws = wb.create_sheet(title="Weekly Summary", index=0)

    ws.append(["WEEKLY STAFFING SUMMARY"])
    ws.cell(1, 1).font = _SZ13
    ws.append([])

    headers = [
        "Day", "Park Open", "LG Start", "LG End", "Park Close",
        "Shore\nShifts", "Reef\nShifts",
        "Adv Coach\n(fixed=2)", "Bay Coach\nShifts",
        "Total\nPersons (PT)",
        "FT-Eligible\nWorkers", "PT\nWorkers",
    ]
    _header_row(ws, headers)

    totals = [0] * 7
    for r in results:
        row = [
            r.day, r.park_open_time, r.lg_start_time, r.lg_end_time, r.park_close_time,
            r.shore_total, r.reef_total, r.ac_total, r.bay_coach_total,
            r.min_persons_all_pt, r.ft_eligible_workers, r.pt_only_workers,
        ]
        ws.append(row)
        for i, v in enumerate([r.shore_total, r.reef_total, r.ac_total,
                                r.bay_coach_total, r.min_persons_all_pt,
                                r.ft_eligible_workers, r.pt_only_workers], 0):
            totals[i] += v

    ws.append([])
    _bold_row(ws, ["WEEK TOTALS", "", "", "", ""] + totals)

    ws.append([])
    ws.append(["EXPLANATION — HOW MIN UNIQUE PERSONS IS CALCULATED"])
    ws.cell(ws.max_row, 1).font = _BOLD
    ws.append([
        "",
        "ALL-PT SCENARIO: Each shift block requires N workers for its time window. "
        "Under all-PT, every worker does exactly one shift — so the minimum unique "
        "person count equals the sum of workers across all shift blocks. "
        "Example (Monday): Shore Guard AM (4 workers × 5.5h) + Shore Guard PM "
        "(4 workers × 7.5h) = 8 Shore persons. Same logic for Reef and Bay.",
    ])
    ws.cell(ws.max_row, 2).alignment = _WRAP

    ws.append([
        "",
        f"FT-ELIGIBLE WORKERS: Any shift ≥ {cfg.FT_SHIFT_THRESHOLD:.0f}h qualifies "
        "as FT-eligible. These workers could be hired full-time (working 5 days/week) "
        f"at ~{cfg.FT_SHIFT_THRESHOLD:.0f}h/day = {cfg.FT_SHIFT_THRESHOLD * 5:.0f}h/week "
        "(exactly the CA FT trigger of 30h/week). The FT count is not necessarily fewer "
        "people per day — it means those positions can be offered to FT employees "
        "rather than requiring separate PT hires for each day.",
    ])
    ws.cell(ws.max_row, 2).alignment = _WRAP

    ws.append([
        "",
        "ADVANCED COACHES: Always 2 per day (1 per shift × 2 shifts), covering the "
        "full LG window. These are separate from Lifeguard headcount. Dual-role "
        "staff (reef-eligible + WSI) may fill an AC shift before or after their "
        "LG shift — they cannot guard and coach at the same time.",
    ])
    ws.cell(ws.max_row, 2).alignment = _WRAP

    ws.append([
        "",
        "30-MIN HANDOVER: Consecutive same-role shifts overlap by exactly 30 min. "
        "The outgoing worker stays 30 min into the next shift for briefing/relief. "
        "Hours shown for each shift include this handover time.",
    ])
    ws.cell(ws.max_row, 2).alignment = _WRAP

    widths = [14, 10, 10, 10, 12, 10, 10, 14, 12, 14, 13, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 35


# ── roster_v0.xlsx — daily schedule grid (LG Schedule style) ─────────────────

def _slot_assignment(t: int, sid: str, sb, slots: list, first_lg_slot: int) -> str | None:
    """
    Return the display value for position `sid` at slot index `t`,
    given that `sid` belongs to shift block `sb`.

      Before shift:  None (blank)
      Arrival slot:  "HH:MM–HH:MM  Prep"  (synthetic for shifts after the first)
      In shift:      rotation station label (Prep / First Aid / Left Shore / …)
      First slot out: "OUT"
      After shift:   None (blank)
    """
    if sb is None:
        return None
    s = sb._start_slot   # first slot covered by this shift
    e = sb._end_slot     # exclusive end

    if t < s - 1:
        return None

    if t == s - 1:
        # Show arrival prep only for shifts that begin after the opening prep slot.
        # The first shift's prep is already embedded in the rotation data.
        if s > first_lg_slot:
            return f"{slots[s].wallclock}–{sb.end}  Prep"
        return None   # let the rotation handle it naturally

    if s <= t < e:
        val = slots[t].assignments.get(sid, "")
        if not val or val == "OFF":
            return None
        if val == "Prep":
            # Annotate with full shift window for the first slot
            return f"{sb.start}–{sb.end}  Prep"
        return val

    if t == e:
        return "OUT"

    return None


def _add_roster_day_sheet(wb, result: DayStaffingResult, slots: list) -> None:
    """
    LG-Schedule-style grid:
      Rows  = 30-min time slots (park open → close)
      Col A = wallclock time
      Remaining columns = one per position, grouped:
        [Shift 1: SG1 SG2 SG3 SF1 | RG1 RG2 RG3 RG4 RF1] [spacer]
        [Shift 2: SG1 SG2 SG3 SF1 | RG1 RG2 RG3 RG4 RF1] [spacer]
        [AC1 AC2] [spacer]
        [BC1 … BC6]

    Two header rows precede the data:
      Row A (desc):   merged shore-rotation description | "Floater" |
                      merged reef-rotation description | spacer | (repeat)
      Row B (labels): shift window | position IDs
    """
    from roster.rotation import staff_layout
    ws = wb.create_sheet(title=result.day)

    layout    = staff_layout()
    shore_rot = [s for s in layout["Shore Guards"]     if s.startswith("SG")]
    shore_flt = [s for s in layout["Shore Guards"]     if s.startswith("SF")]
    reef_rot  = [s for s in layout["Reef Guards"]      if s.startswith("RG")]
    reef_flt  = [s for s in layout["Reef Guards"]      if s.startswith("RF")]
    ac_ids    = layout["Advanced Coaches"]
    bay_ids   = layout["Bay Coaches"]

    # Sort shift blocks by start slot
    role_shifts: dict[str, list] = {}
    for sb in result.shifts:
        role_shifts.setdefault(sb.role, []).append(sb)
    for v in role_shifts.values():
        v.sort(key=lambda b: b._start_slot)

    shore_blocks = role_shifts.get("Shore Guard",    [])
    reef_blocks  = role_shifts.get("Reef Guard",     [])
    ac_blocks    = role_shifts.get("Advanced Coach", [])
    bay_blocks   = role_shifts.get("Bay Coach",      [])

    # first_lg_slot: used to decide whether to show synthetic Prep
    from solver.staffing import _demand_vectors
    demand = _demand_vectors(slots)
    lg_demand_slots = [t for t in range(len(slots))
                       if demand["shore"][t] + demand["reef"][t] > 0]
    first_lg_slot = lg_demand_slots[0] if lg_demand_slots else 0

    # ── Build flat column list ─────────────────────────────────────────────────
    # Each entry: (sid, shift_block, fill_color) or None (spacer)
    col_list: list = []
    n_lg_shifts = max(len(shore_blocks), len(reef_blocks), 1)

    for i in range(n_lg_shifts):
        sb_s = shore_blocks[i] if i < len(shore_blocks) else None
        sb_r = reef_blocks[i]  if i < len(reef_blocks)  else None
        for sid in shore_rot:
            col_list.append((sid, sb_s, _BLUE))
        for sid in shore_flt:
            col_list.append((sid, sb_s, _TEAL))
        for sid in reef_rot:
            col_list.append((sid, sb_r, _AMBER))
        for sid in reef_flt:
            col_list.append((sid, sb_r, _ROSE))
        col_list.append(None)   # spacer

    for j, ac_sb in enumerate(ac_blocks):
        sid = ac_ids[j] if j < len(ac_ids) else f"AC{j+1}"
        col_list.append((sid, ac_sb, _GREEN))
    if ac_blocks:
        col_list.append(None)

    bay_sb = bay_blocks[0] if bay_blocks else None
    for sid in bay_ids:
        col_list.append((sid, bay_sb, _LILAC))

    # ── Title row ──────────────────────────────────────────────────────────────
    n_data_cols = 1 + len(col_list)
    ws.append([
        f"{result.day}  ·  Park {result.park_open_time}–{result.park_close_time}"
        f"  ·  LG {result.lg_start_time}–{result.lg_end_time}"
    ])
    ws.cell(1, 1).font = _SZ13
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=min(n_data_cols, 60))

    # ── Header row A: group/rotation descriptions ──────────────────────────────
    ws.append([None] + [None] * len(col_list))
    desc_row = ws.max_row

    def _merge_fill(r, c1, c2, text, fill, font=None):
        ws.cell(r, c1).value = text
        ws.cell(r, c1).fill = fill
        ws.cell(r, c1).alignment = _WCTR
        if font:
            ws.cell(r, c1).font = font
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1,
                           end_row=r, end_column=c2)
            for c in range(c1 + 1, c2 + 1):
                ws.cell(r, c).fill = fill

    col_cursor = 2   # 1-indexed; col 1 = time
    for i in range(n_lg_shifts):
        sb_s = shore_blocks[i] if i < len(shore_blocks) else None
        sb_r = reef_blocks[i]  if i < len(reef_blocks)  else None
        window = (sb_s or sb_r)
        shift_label = (f"SHIFT {i+1}  {window.start}–{window.end}"
                       if window else f"SHIFT {i+1}")

        c_shore_start = col_cursor
        c_shore_end   = col_cursor + len(shore_rot) - 1
        c_flt         = c_shore_end + 1
        c_reef_start  = c_flt + len(shore_flt)
        c_reef_end    = c_reef_start + len(reef_rot) - 1
        c_rflt_end    = c_reef_end + len(reef_flt)

        _merge_fill(desc_row, c_shore_start, c_shore_end,
                    f"Shore Guard — FA → Left Shore → Right Shore  (rotation hourly)\n{shift_label}",
                    _BLUE, _BOLD)
        for k, _ in enumerate(shore_flt):
            _merge_fill(desc_row, c_flt + k, c_flt + k, "Floater", _TEAL)
        _merge_fill(desc_row, c_reef_start, c_reef_end,
                    f"Reef Guard — Rental → Tower → Left Reef → Right Reef  (rotation hourly)\n{shift_label}",
                    _AMBER, _BOLD)
        for k, _ in enumerate(reef_flt):
            _merge_fill(desc_row, c_reef_end + 1 + k, c_reef_end + 1 + k,
                        "Reef Floater", _ROSE)
        col_cursor = c_rflt_end + 1 + 1   # +1 for spacer col

    # AC descriptions
    for j, ac_sb in enumerate(ac_blocks):
        sid = ac_ids[j] if j < len(ac_ids) else f"AC{j+1}"
        sh_label = f"{ac_sb.start}–{ac_sb.end}" if ac_sb else ""
        _merge_fill(desc_row, col_cursor, col_cursor,
                    f"Advanced Coach\n{sh_label}", _GREEN, _BOLD)
        col_cursor += 1
    if ac_blocks:
        col_cursor += 1   # spacer

    # Bay description
    if bay_ids and bay_sb:
        c_bay_end = col_cursor + len(bay_ids) - 1
        sh_label  = f"{bay_sb.start}–{bay_sb.end}"
        _merge_fill(desc_row, col_cursor, c_bay_end,
                    f"Bay Coach\n{sh_label}", _LILAC, _BOLD)

    ws.row_dimensions[desc_row].height = 36

    # ── Header row B: shift window + position IDs ──────────────────────────────
    row_b = [f"{result.park_open_time}–{result.park_close_time}"]
    for entry in col_list:
        if entry is None:
            row_b.append(None)
        else:
            sid, _, _ = entry
            row_b.append(sid)
    ws.append(row_b)
    label_row = ws.max_row
    ws.cell(label_row, 1).font = _BOLD
    ws.cell(label_row, 1).fill = _GRAY
    ws.cell(label_row, 1).alignment = _CTR
    for c, entry in enumerate(col_list, 2):
        cell = ws.cell(label_row, c)
        cell.font = _BOLD
        cell.alignment = _CTR
        if entry is not None:
            _, _, fill = entry
            cell.fill = fill
        else:
            cell.fill = _GRAY

    ws.freeze_panes = f"B{label_row + 1}"

    # ── Data rows: one per 30-min slot ────────────────────────────────────────
    for t, slot in enumerate(slots):
        row = [slot.wallclock]
        for entry in col_list:
            if entry is None:
                row.append(None)
            else:
                sid, sb, _ = entry
                row.append(_slot_assignment(t, sid, sb, slots, first_lg_slot))
        ws.append(row)
        data_row = ws.max_row

        # Style: bold time, colour filled cells
        ws.cell(data_row, 1).font  = _BOLD
        ws.cell(data_row, 1).alignment = _CTR
        for c, entry in enumerate(col_list, 2):
            cell = ws.cell(data_row, c)
            cell.alignment = _CTR
            if entry is not None and cell.value is not None:
                _, _, fill = entry
                cell.fill = fill
                if cell.value in ("OUT",):
                    cell.font = Font(bold=True, color="FF666666")
            elif cell.value == "OUT":
                cell.font = Font(bold=True, color="FF666666")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7
    for c, entry in enumerate(col_list, 2):
        col_letter = get_column_letter(c)
        if entry is None:
            ws.column_dimensions[col_letter].width = 2
        else:
            ws.column_dimensions[col_letter].width = 16

    _bold_row(ws, [])   # trailing blank row for visual spacing


# ── roster_v0.xlsx — shift template summary ───────────────────────────────────

def _add_shift_template_sheet(wb, results: list[DayStaffingResult]) -> None:
    """
    One table per distinct operating-day type showing the shift pattern
    (start/end times per role) that applies to those days.
    """
    ws = wb.create_sheet(title="Shift Templates", index=0)

    ws.append(["SHIFT TEMPLATES BY OPERATING DAY TYPE"])
    ws.cell(1, 1).font = _SZ13
    ws.append([
        "Each table below shows the shift structure for one class of operating day. "
        "Days with the same LG window length and park hours share a template."
    ])
    ws.cell(ws.max_row, 1).alignment = _WRAP
    ws.append([])

    # Group days by (park_open, park_close) → representative template
    from collections import OrderedDict
    seen: dict[tuple, list[str]] = OrderedDict()
    template_shifts: dict[tuple, list[ShiftBlock]] = {}
    for r in results:
        key = (r.park_open_time, r.park_close_time, r.lg_start_time, r.lg_end_time)
        seen.setdefault(key, []).append(r.day)
        template_shifts[key] = r.shifts  # last one wins; all same-key days are identical

    for key, days in seen.items():
        open_t, close_t, lg_start, lg_end = key
        lg_mins = sum(1 for _ in range(0)) or (
            int(lg_end.split(":")[0]) * 60 + int(lg_end.split(":")[1])
            - int(lg_start.split(":")[0]) * 60 - int(lg_start.split(":")[1])
        )
        day_list = ", ".join(days)
        ws.append([f"Applies to: {day_list}"])
        ws.cell(ws.max_row, 1).font = _BOLD
        ws.append([
            f"Park {open_t}–{close_t}   ·   LG window {lg_start}–{lg_end} "
            f"({lg_mins // 60}h {lg_mins % 60:02d}min)"
        ])
        ws.append([])

        _header_row(ws, [
            "Role", "Shift", "Start", "End",
            "Workers\nper Shift", "Hours\n/Worker", "CA Break",
        ])

        shifts = template_shifts[key]
        role_seen: dict[str, int] = {}
        for sb in shifts:
            role_seen.setdefault(sb.role, 0)
            role_seen[sb.role] += 1
            shift_n = role_seen[sb.role]
            ft = f"FT-eligible (≥{cfg.FT_SHIFT_THRESHOLD:.0f}h)" if sb.shift_hrs >= cfg.FT_SHIFT_THRESHOLD else "PT"
            ws.append([
                sb.role, f"Shift {shift_n}", sb.start, sb.end,
                sb.workers, sb.shift_hrs, sb.ca_note or "—",
            ])
            fill = _role_fill(sb.role)
            for col in range(1, 8):
                ws.cell(ws.max_row, col).fill = fill
            ws.cell(ws.max_row, 7).alignment = _WRAP

        ws.append([])
        ws.append([])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 52
    ws.row_dimensions[1].height = 20


# ── Public render functions ───────────────────────────────────────────────────

def render_staffing_xlsx(
    results: list[DayStaffingResult],
    out_path: str | Path,
) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_weekly_summary(wb, results)
    for r in results:
        _add_day_sheet(wb, r)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def render_roster_xlsx(
    results: list[DayStaffingResult],
    slots_by_day: dict[str, list],
    out_path: str | Path,
) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_shift_template_sheet(wb, results)
    for r in results:
        _add_roster_day_sheet(wb, r, slots_by_day[r.day])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
